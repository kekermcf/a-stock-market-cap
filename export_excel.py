import json, os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

with open(f'{DATA_DIR}/stock_data_full.json', 'r', encoding='utf-8') as f:
    results = json.load(f)

with open(f'{DATA_DIR}/cache/ah_status.json', 'r', encoding='utf-8') as f:
    ah_status = json.load(f)

with open(f'{DATA_DIR}/cache/sanction_list.json', 'r', encoding='utf-8') as f:
    sanction_list = json.load(f)

with open(f'{DATA_DIR}/cache/ah_hk_list_dates.json', 'r', encoding='utf-8') as f:
    ah_hk_dates = json.load(f)

with open(f'{DATA_DIR}/cache/ah_premia.json', 'r', encoding='utf-8') as f:
    ah_premia = json.load(f)

def get_sanction_label(ts_code):
    if ts_code not in sanction_list:
        return ''
    labels = sanction_list[ts_code]
    if 'NS-CMIC' in labels and 'Entity List' in labels:
        return 'NS-CMIC+EL'
    if 'Entity List' in labels:
        return 'Entity List'
    if 'NS-CMIC' in labels:
        return 'NS-CMIC'
    return ''

headers = [
    '排名', '代码', '名称', '行业', '主营业务',
    'A+H状态', '制裁', 'H上市日', 'H/A溢价',
    '总市值(亿)', '24TD', '25TD', 'YTD',
    '2025收入(亿)', '毛利率', '净利润(亿)', '净利率'
]

rows = []
for r in results:
    ts = r['ts_code']
    ah = ah_status.get(ts, '')

    if ah == 'listed':
        ah_tag = 'A+H'
    elif ah == 'announced':
        ah_tag = '拟H股'
    elif ah == 'rumor':
        ah_tag = 'Rumor'
    else:
        ah_tag = ''

    sc_label = get_sanction_label(ts)

    prem = ah_premia.get(ts)
    if prem is not None and ah == 'listed':
        h_a_prem = -prem
        prem_str = f"H/A:{'+' if h_a_prem >= 0 else ''}{h_a_prem:.1f}%"
    else:
        prem_str = ''

    ytd24 = r.get('ytd_2024')
    ytd25 = r.get('ytd_2025')
    ytd26 = r.get('ytd_2026')

    row = [
        r.get('_rank', ''),
        ts,
        r.get('name', ''),
        r.get('industry', ''),
        r.get('business_desc', ''),
        ah_tag,
        sc_label,
        ah_hk_dates.get(ts, '') if ah == 'listed' else '',
        prem_str if ah == 'listed' else '',
        round(r.get('total_mv', 0), 1),
        ytd24 if ytd24 is not None else '',
        ytd25 if ytd25 is not None else '',
        ytd26 if ytd26 is not None else '',
        r.get('revenue'),
        f"{r['gpr']:.1f}%" if r.get('gpr') is not None else '',
        r.get('net_profit'),
        f"{r['npm']:.1f}%" if r.get('npm') is not None else '',
    ]
    rows.append(row)

# Sort by total_mv descending
rows.sort(key=lambda x: float(x[9]) if x[9] else 0, reverse=True)
for i, row in enumerate(rows):
    row[0] = i + 1

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = 'A股超200亿市值'

header_font = Font(bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='FF0d9488', end_color='FF0d9488', fill_type='solid')
thin = Side(style='thin')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

center_cols = [1, 6, 7, 8]  # rank, A+H状态, 制裁, H上市日

for row_idx, row in enumerate(rows, 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        if col_idx in center_cols:
            cell.alignment = Alignment(horizontal='center')
        elif isinstance(val, (int, float)) and col_idx not in [2, 3, 4, 5]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

col_widths = [6, 12, 10, 12, 40, 10, 14, 12, 14, 12, 8, 8, 8, 14, 8, 14, 8]
for col_idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'

out_path = f'{DATA_DIR}/A股超200亿市值.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Total rows: {len(rows)}')
